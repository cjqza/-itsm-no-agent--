"""OPS API - 查询与统计"""
from fastapi import APIRouter, Depends, Query
from sqlalchemy import select, func, and_, case, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from typing import Optional
from datetime import datetime, timedelta, timezone

from app.database import get_db
from app.models.user import User, UserRole
from app.models.ticket import Ticket, TicketStatus, SLAStatus
from app.models.category import Category
from app.utils.auth import require_permission
from app.utils import escape_like

router = APIRouter(prefix="/api/ops", tags=["OPS"])


@router.get("/statistics/overview")
async def statistics_overview(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """总览统计"""
    since = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime(2000, 1, 1, tzinfo=timezone.utc)

    # 总工单数
    total_result = await db.execute(
        select(func.count(Ticket.id)).where(Ticket.created_at >= since)
    )

    # 按状态统计（单次查询）
    status_result = await db.execute(
        select(
            Ticket.status,
            func.count(Ticket.id).label("count"),
        )
        .where(Ticket.created_at >= since)
        .group_by(Ticket.status)
    )
    status_counts = {row.status.value: row.count for row in status_result.all()}
    # 补全缺失的状态
    for s in TicketStatus:
        status_counts.setdefault(s.value, 0)

    # 平均评分
    avg_result = await db.execute(
        select(func.avg(Ticket.rating)).where(
            Ticket.rating.isnot(None),
            Ticket.created_at >= since,
        )
    )

    # SLA达标率（单次查询）
    sla_result = await db.execute(
        select(
            func.count(Ticket.id).label("total_resolved"),
            func.sum(
                case(
                    (Ticket.sla_status.in_([SLAStatus.GREEN, SLAStatus.YELLOW]), 1),
                    else_=0,
                )
            ).label("sla_met"),
        ).where(
            Ticket.status == TicketStatus.RESOLVED,
            Ticket.created_at >= since,
        )
    )
    sla_row = sla_result.one()
    total_resolved = sla_row.total_resolved or 0
    sla_met = sla_row.sla_met or 0
    sla_rate = (sla_met / total_resolved * 100) if total_resolved > 0 else 100.0

    return {
        "total": total_result.scalar() or 0,
        "status_counts": status_counts,
        "avg_rating": round(float(avg_result.scalar() or 0), 2),
        "sla_compliance_rate": round(sla_rate, 2),
    }


@router.get("/statistics/by-category")
async def statistics_by_category(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """按管理单元统计"""
    since = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime(2000, 1, 1, tzinfo=timezone.utc)

    result = await db.execute(
        select(
            Category.name,
            func.count(Ticket.id).label("count"),
        )
        .outerjoin(
            Ticket,
            and_(Ticket.category_id == Category.id, Ticket.created_at >= since),
        )
        .group_by(Category.id, Category.name)
        .order_by(func.count(Ticket.id).desc())
    )

    return [{"name": row[0], "count": row[1]} for row in result.all()]


@router.get("/statistics/by-agent")
async def statistics_by_agent(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """按客服统计"""
    since = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime(2000, 1, 1, tzinfo=timezone.utc)

    result = await db.execute(
        select(
            User.name,
            func.count(Ticket.id).label("total"),
            func.sum(
                case(
                    (Ticket.status == TicketStatus.RESOLVED, 1),
                    else_=0,
                )
            ).label("resolved"),
            func.avg(Ticket.rating).label("avg_rating"),
        )
        .outerjoin(
            Ticket,
            and_(Ticket.assignee_id == User.id, Ticket.created_at >= since),
        )
        .where(User.role == UserRole.AGENT)
        .group_by(User.id, User.name)
        .order_by(func.count(Ticket.id).desc())
    )

    return [
        {
            "name": row[0],
            "total": row[1],
            "resolved": int(row[2] or 0),
            "avg_rating": round(float(row[3] or 0), 2),
        }
        for row in result.all()
    ]


@router.get("/statistics/ratings")
async def statistics_ratings(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """评价统计"""
    since = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime(2000, 1, 1, tzinfo=timezone.utc)

    # 评分分布（单次查询）
    dist_result = await db.execute(
        select(
            Ticket.rating,
            func.count(Ticket.id).label("count"),
        )
        .where(
            Ticket.rating.isnot(None),
            Ticket.created_at >= since,
        )
        .group_by(Ticket.rating)
    )
    dist_map = {row.rating: row.count for row in dist_result.all()}
    distribution = [{"rating": i, "count": dist_map.get(i, 0)} for i in range(1, 6)]

    # 最近评价
    recent = await db.execute(
        select(Ticket)
        .where(
            Ticket.rating.isnot(None),
            Ticket.created_at >= since,
        )
        .order_by(Ticket.rated_at.desc())
        .limit(10)
    )

    recent_ratings = [
        {
            "ticket_no": t.ticket_no,
            "title": t.title,
            "rating": t.rating,
            "comment": t.rating_comment,
            "rated_at": t.rated_at.isoformat() if t.rated_at else None,
        }
        for t in recent.scalars().all()
    ]

    return {
        "distribution": distribution,
        "recent": recent_ratings,
    }


@router.get("/statistics/sla-compliance")
async def sla_compliance(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """SLA达标率详情"""
    since = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime(2000, 1, 1, tzinfo=timezone.utc)

    result = await db.execute(
        select(
            Category.name,
            func.count(Ticket.id).label("total"),
            func.sum(
                case(
                    (Ticket.sla_status.in_([SLAStatus.GREEN, SLAStatus.YELLOW]), 1),
                    else_=0,
                )
            ).label("met"),
        )
        .outerjoin(
            Ticket,
            and_(Ticket.category_id == Category.id, Ticket.created_at >= since),
        )
        .group_by(Category.id, Category.name)
    )

    return [
        {
            "category": row[0] or "未分类",
            "total": row[1],
            "met": int(row[2] or 0),
            "rate": round(int(row[2] or 0) / row[1] * 100, 2) if row[1] > 0 else 100,
        }
        for row in result.all()
    ]


@router.get("/statistics/trend")
async def statistics_trend(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """趋势分析（按天）"""
    since = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime(2000, 1, 1, tzinfo=timezone.utc)

    result = await db.execute(
        select(
            func.date(Ticket.created_at).label("date"),
            func.count().label("count"),
        )
        .where(Ticket.created_at >= since)
        .group_by(func.date(Ticket.created_at))
        .order_by(func.date(Ticket.created_at))
    )

    return [{"date": str(row.date), "count": row.count} for row in result.all()]


@router.get("/export")
async def export_tickets(
    days: Optional[int] = Query(None, ge=1, le=365),
    status: Optional[str] = None,
    category_id: Optional[int] = None,
    keyword: Optional[str] = None,
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """导出工单报表（支持与列表一致的筛选：状态/管理单元/关键字）"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    import io

    since = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime(2000, 1, 1, tzinfo=timezone.utc)

    conditions = [Ticket.created_at >= since]
    if status:
        conditions.append(Ticket.status == status)
    if category_id:
        conditions.append(Ticket.category_id == category_id)
    if keyword:
        safe_kw = escape_like(keyword)
        conditions.append(
            or_(
                Ticket.ticket_no.like(f"%{safe_kw}%", escape="\\"),
                Ticket.title.like(f"%{safe_kw}%", escape="\\"),
            )
        )

    result = await db.execute(
        select(Ticket)
        .options(selectinload(Ticket.category))
        .where(*conditions)
        .order_by(Ticket.created_at.desc())
    )
    tickets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "工单报表"

    headers = [
        "工单号", "标题", "状态", "优先级", "管理单元",
        "SLA状态", "评分", "创建时间", "接单时间", "解决时间",
    ]
    ws.append(headers)

    # 设置表头样式
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    status_map = {
        "pending": "待派发", "assigned": "已派发", "accepted": "已接单",
        "analyzing": "分析中", "processing": "处理中",
        "resolved_pending_review": "待评价", "resolved": "已解决",
    }

    for t in tickets:
        ws.append([
            t.ticket_no,
            t.title,
            status_map.get(t.status.value, t.status.value) if t.status else "",
            t.priority.value if t.priority else "",
            t.category.name if t.category else "",
            t.sla_status.value if t.sla_status else "",
            t.rating or "",
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
            t.accepted_at.strftime("%Y-%m-%d %H:%M") if t.accepted_at else "",
            t.resolved_at.strftime("%Y-%m-%d %H:%M") if t.resolved_at else "",
        ])

    # 自动调整列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tickets_report.xlsx"},
    )


# ==================== 新增端点 ====================


@router.get("/tickets")
async def list_tickets(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    category_id: Optional[int] = None,
    keyword: Optional[str] = None,
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """OPS工单列表（查看所有工单）"""
    from app.services.ticket_service import ticket_service

    query = select(Ticket).options(
        selectinload(Ticket.category),
        selectinload(Ticket.creator),
        selectinload(Ticket.assignee),
    )

    if status:
        query = query.where(Ticket.status == status)
    if category_id:
        query = query.where(Ticket.category_id == category_id)
    if keyword:
        safe_kw = escape_like(keyword)
        query = query.where(
            or_(
                Ticket.ticket_no.like(f"%{safe_kw}%", escape="\\"),
                Ticket.title.like(f"%{safe_kw}%", escape="\\"),
            )
        )

    # Count
    count_query = select(func.count(Ticket.id))
    if status:
        count_query = count_query.where(Ticket.status == status)
    if category_id:
        count_query = count_query.where(Ticket.category_id == category_id)
    if keyword:
        safe_kw = escape_like(keyword)
        count_query = count_query.where(
            or_(
                Ticket.ticket_no.like(f"%{safe_kw}%", escape="\\"),
                Ticket.title.like(f"%{safe_kw}%", escape="\\"),
            )
        )
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    query = query.order_by(Ticket.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    tickets = result.scalars().all()

    items = []
    for t in tickets:
        d = ticket_service._ticket_to_dict(t)
        d["category_name"] = t.category.name if t.category else None
        d["creator_name"] = t.creator.name if t.creator else None
        d["assignee_name"] = t.assignee.name if t.assignee else None
        items.append(d)

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/status-distribution")
async def status_distribution(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """工单状态分布"""
    conditions = []
    if days:
        since = datetime.now(timezone.utc) - timedelta(days=days)
        conditions.append(Ticket.created_at >= since)

    result = await db.execute(
        select(Ticket.status, func.count(Ticket.id).label("count"))
        .where(*conditions)
        .group_by(Ticket.status)
    )
    status_counts = {row.status.value if hasattr(row.status, 'value') else row.status: row.count for row in result.all()}
    for s in TicketStatus:
        status_counts.setdefault(s.value, 0)
    return status_counts


@router.get("/category-stats")
async def category_stats(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """管理单元统计（含平均处理时长）"""
    conditions = []
    if days:
        since = datetime.now(timezone.utc) - timedelta(days=days)
        conditions.append(Ticket.created_at >= since)

    result = await db.execute(
        select(
            Category.name,
            func.count(Ticket.id).label("count"),
            func.avg(
                case(
                    (
                        and_(Ticket.resolved_at.isnot(None), Ticket.accepted_at.isnot(None)),
                        func.julianday(Ticket.resolved_at) - func.julianday(Ticket.accepted_at),
                    ),
                    else_=None,
                )
            ).label("avg_hours"),
        )
        .outerjoin(
            Ticket,
            and_(Ticket.category_id == Category.id, *conditions),
        )
        .group_by(Category.id, Category.name)
        .order_by(func.count(Ticket.id).desc())
    )

    return [
        {
            "category_name": row[0] or "未分类",
            "count": row[1],
            "avg_hours": round(float(row[2]) * 24, 1) if row[2] else 0,
        }
        for row in result.all()
    ]


@router.get("/rating-distribution")
async def rating_distribution(
    days: Optional[int] = Query(None, ge=1, le=365),
    current_user: User = Depends(require_permission("ops_access")),
    db: AsyncSession = Depends(get_db),
):
    """评分分布"""
    conditions = [Ticket.rating.isnot(None)]
    if days:
        since = datetime.now(timezone.utc) - timedelta(days=days)
        conditions.append(Ticket.created_at >= since)

    result = await db.execute(
        select(Ticket.rating, func.count(Ticket.id).label("count"))
        .where(*conditions)
        .group_by(Ticket.rating)
    )
    dist_map = {row.rating: row.count for row in result.all()}
    return {f"rating_{i}": dist_map.get(i, 0) for i in range(1, 6)}
